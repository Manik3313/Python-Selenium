import  openpyxl
class getdata:
    @staticmethod
    def getdatafromexcel(path,name):
        book = openpyxl.load_workbook(path)
        sheet = book.active#get_sheet_by_name(name)
        var = []
        for i in range(2, sheet.max_row + 1):  # to get rows
            # if sheet.cell(row=i, column=1).value == test_case_name:
            Dict = {}
            for j in range(1, sheet.max_column + 1):  # to get columns
                # Dict["lastname"]="shetty
                Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
            print(Dict)
            var.insert(i - 1, Dict)
        return var

    """"
        for i in range(1, sheet.max_row + 1):  # to get rows
            if sheet.cell(row=i, column=1).value == test_case_name:

                for j in range(2, sheet.max_column + 1):  # to get columns
                    # Dict["lastname"]="shetty
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
        return [Dict]
    """""